"""
API endpoints cho hệ thống bán hải sản
"""
from ninja import Router
from pydantic import BaseModel
from typing import List
from django.shortcuts import get_object_or_404
from django.db.models import Sum, F, Q
from django.utils import timezone
from decimal import Decimal
from uuid import UUID
import uuid

from .models import (
    SeafoodCategory, Seafood, ImportSource, ImportBatch,
    Order, OrderItem, InventoryLog
)
from .schemas import (
    CategoryRead, CategoryCreate, CategoryUpdate,
    SeafoodRead, SeafoodCreate, SeafoodUpdate,
    ImportSourceRead, ImportSourceCreate, ImportSourceUpdate,
    ImportBatchRead, ImportBatchCreate, ImportBatchUpdate,
    OrderRead, OrderCreate, OrderUpdate, OrderItemRead,
    OrderConfirmBySale, OrderAssignToEmployee, OrderStartWeighing, OrderCompleteWeighing,
    DashboardStats, ProductStats
)
from .sepay_service import get_sepay_service, SepayAPIError

router = Router(tags=["Seafood"], auth=None)  # No auth required for seafood APIs


# ============================================
# CATEGORY ENDPOINTS
# ============================================

@router.get("/categories", response=List[CategoryRead])
def list_categories(request):
    """Lấy danh sách danh mục"""
    return SeafoodCategory.objects.filter(is_active=True).all()


@router.post("/categories", response=CategoryRead)
def create_category(request, payload: CategoryCreate):
    """Tạo danh mục mới"""
    category = SeafoodCategory.objects.create(**payload.dict())
    return category


@router.put("/categories/{category_id}", response=CategoryRead)
def update_category(request, category_id: UUID, payload: CategoryUpdate):
    """Cập nhật danh mục"""
    category = get_object_or_404(SeafoodCategory, id=category_id)
    for key, value in payload.dict(exclude_unset=True).items():
        setattr(category, key, value)
    category.save()
    return category


@router.delete("/categories/{category_id}")
def delete_category(request, category_id: UUID):
    """Xóa danh mục (soft delete)"""
    category = get_object_or_404(SeafoodCategory, id=category_id)
    category.is_active = False
    category.save()
    return {"success": True}


# ============================================
# SEAFOOD PRODUCTS ENDPOINTS
# ============================================

@router.get("/products", response=List[SeafoodRead])
def list_products(request, category_id: UUID = None, status: str = None, search: str = None):
    """Lấy danh sách sản phẩm hải sản"""
    query = Seafood.objects.filter(is_active=True).select_related('category')

    if category_id:
        query = query.filter(category_id=category_id)
    if status:
        query = query.filter(status=status)
    if search:
        query = query.filter(Q(name__icontains=search) | Q(code__icontains=search))

    return query.all()


# ============================================
# IMPORT/EXPORT ENDPOINTS (must be before /{product_id})
# ============================================

@router.post("/products/import-excel")
def import_products_excel(request):
    """Import sản phẩm từ file Excel"""
    from django.http import JsonResponse
    from openpyxl import load_workbook
    from django.utils.text import slugify

    # Get uploaded file
    excel_file = request.FILES.get('file')
    if not excel_file:
        return JsonResponse({"error": "No file uploaded"}, status=400)

    try:
        # Load workbook
        wb = load_workbook(excel_file)
        ws = wb.active

        created_count = 0
        updated_count = 0
        errors = []

        # Skip header row (row 1)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Expected columns: Mã, Tên, Danh mục, Đơn vị, Giá, Tồn kho, Xuất xứ, Trạng thái
                if not row[0] or not row[1]:  # Skip if no code or name
                    continue

                code = str(row[0]).strip()
                name = str(row[1]).strip()
                category_name = str(row[2]).strip() if row[2] else None
                unit_type = str(row[3]).strip().lower() if row[3] else 'kg'
                current_price = float(row[4]) if row[4] else 0
                stock_quantity = float(row[5]) if row[5] else 0
                origin = str(row[6]).strip() if row[6] else ''
                status = str(row[7]).strip().lower() if row[7] else 'active'

                # Map unit type
                unit_type_map = {
                    'kg': 'kg',
                    'kilogram': 'kg',
                    'con': 'piece',
                    'cái': 'piece',
                    'piece': 'piece',
                    'thùng': 'box',
                    'hộp': 'box',
                    'box': 'box',
                }
                unit_type = unit_type_map.get(unit_type, 'kg')

                # Map status
                status_map = {
                    'active': 'active',
                    'đang bán': 'active',
                    'hoạt động': 'active',
                    'inactive': 'inactive',
                    'ngừng bán': 'inactive',
                    'ngưng bán': 'inactive',
                }
                status = status_map.get(status, 'active')

                # Find or create category
                category = None
                if category_name:
                    category, _ = SeafoodCategory.objects.get_or_create(
                        name=category_name,
                        defaults={
                            'slug': slugify(category_name),
                            'description': f'Danh mục {category_name}',
                        }
                    )

                # Create or update product
                product, created = Seafood.objects.update_or_create(
                    code=code,
                    defaults={
                        'name': name,
                        'category': category,
                        'unit_type': unit_type,
                        'current_price': Decimal(str(current_price)),
                        'stock_quantity': Decimal(str(stock_quantity)),
                        'origin': origin,
                        'status': status,
                    }
                )

                if created:
                    created_count += 1
                else:
                    updated_count += 1

            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                continue

        return {
            "success": True,
            "created": created_count,
            "updated": updated_count,
            "errors": errors
        }

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


@router.get("/products/export-excel")
def export_products_excel(request):
    """Export danh sách sản phẩm ra Excel"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    # Get all products
    products = Seafood.objects.filter(is_active=True).select_related('category').order_by('code')

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sản phẩm"

    # Header style
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = ['Mã sản phẩm', 'Tên sản phẩm', 'Danh mục', 'Đơn vị', 'Giá (đ)', 'Tồn kho', 'Xuất xứ', 'Trạng thái']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Data rows
    for row_idx, product in enumerate(products, start=2):
        ws.cell(row=row_idx, column=1, value=product.code)
        ws.cell(row=row_idx, column=2, value=product.name)
        ws.cell(row=row_idx, column=3, value=product.category.name if product.category else '')

        # Unit type display
        unit_type_display = {
            'kg': 'kg',
            'piece': 'Con/Cái',
            'box': 'Thùng/Hộp'
        }
        ws.cell(row=row_idx, column=4, value=unit_type_display.get(product.unit_type, product.unit_type))

        ws.cell(row=row_idx, column=5, value=float(product.current_price))
        ws.cell(row=row_idx, column=6, value=float(product.stock_quantity))
        ws.cell(row=row_idx, column=7, value=product.origin or '')

        # Status display
        status_display = {
            'active': 'Đang bán',
            'inactive': 'Ngừng bán',
            'out_of_stock': 'Hết hàng'
        }
        ws.cell(row=row_idx, column=8, value=status_display.get(product.status, product.status))

    # Adjust column widths
    column_widths = [15, 30, 20, 12, 15, 12, 20, 15]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Return response
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="san-pham-{timezone.now().strftime("%Y%m%d")}.xlsx"'

    return response


@router.get("/products/export-pdf")
def export_products_pdf(request):
    """Export bảng giá sản phẩm ra PDF"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO

    # Get all active products
    products = Seafood.objects.filter(is_active=True, status='active').select_related('category').order_by('category__name', 'code')

    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []

    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    elements.append(Paragraph("BẢNG GIÁ SẢN PHẨM HẢI SẢN", title_style))
    elements.append(Paragraph(f"Ngày cập nhật: {timezone.now().strftime('%d/%m/%Y')}",
                             ParagraphStyle('Subtitle', parent=styles['Normal'], alignment=TA_CENTER, fontSize=10)))
    elements.append(Spacer(1, 15))

    # Group products by category
    category_groups = {}
    for product in products:
        cat_name = product.category.name if product.category else 'Chưa phân loại'
        if cat_name not in category_groups:
            category_groups[cat_name] = []
        category_groups[cat_name].append(product)

    # Create table for each category
    for category_name, cat_products in category_groups.items():
        # Category header
        cat_style = ParagraphStyle('CategoryHeader', parent=styles['Heading2'],
                                   fontSize=14, textColor=colors.HexColor('#059669'),
                                   spaceAfter=10, spaceBefore=10)
        elements.append(Paragraph(category_name, cat_style))

        # Table data
        table_data = [['Mã SP', 'Tên sản phẩm', 'Đơn vị', 'Giá (đ)', 'Tồn kho']]

        for product in cat_products:
            unit_display = {'kg': 'kg', 'piece': 'Con', 'box': 'Thùng'}.get(product.unit_type, product.unit_type)
            table_data.append([
                product.code,
                product.name,
                unit_display,
                f"{product.current_price:,.0f}",
                f"{product.stock_quantity:,.1f}"
            ])

        # Create table
        table = Table(table_data, colWidths=[4*cm, 10*cm, 3*cm, 4*cm, 3*cm])
        table.setStyle(TableStyle([
            # Header style
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),

            # Data style
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Code
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # Name
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),  # Unit
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),   # Price
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),   # Stock
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 15))

    # Build PDF
    doc.build(elements)

    # Return response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="bang-gia-san-pham-{timezone.now().strftime("%Y%m%d")}.pdf"'

    return response


# ============================================
# PRODUCT DETAIL ENDPOINTS
# ============================================

@router.get("/products/{product_id}", response=SeafoodRead)
def get_product(request, product_id: UUID):
    """Lấy chi tiết sản phẩm"""
    return get_object_or_404(Seafood.objects.select_related('category'), id=product_id)


@router.post("/products", response=SeafoodRead)
def create_product(request, payload: SeafoodCreate):
    """Tạo sản phẩm mới"""
    product = Seafood.objects.create(**payload.dict())
    return product


@router.put("/products/{product_id}", response=SeafoodRead)
def update_product(request, product_id: UUID, payload: SeafoodUpdate):
    """Cập nhật sản phẩm"""
    product = get_object_or_404(Seafood, id=product_id)
    for key, value in payload.dict(exclude_unset=True).items():
        setattr(product, key, value)
    product.save()
    return product


@router.delete("/products/{product_id}")
def delete_product(request, product_id: UUID):
    """Xóa sản phẩm (soft delete)"""
    product = get_object_or_404(Seafood, id=product_id)
    product.is_active = False
    product.save()
    return {"success": True}


# ============================================
# INVENTORY MANAGEMENT ENDPOINTS
# ============================================

@router.post("/products/{product_id}/adjust-inventory")
def adjust_inventory(request, product_id: UUID):
    """
    Điều chỉnh tồn kho sản phẩm và tự động ghi log
    Loại điều chỉnh: import (nhập hàng), adjust (điều chỉnh), loss (hao hụt)
    """
    from django.http import JsonResponse
    import json

    try:
        # Get request body
        body = json.loads(request.body)
        adjustment_type = body.get('type', 'adjust')  # import, adjust, loss
        quantity = Decimal(str(body.get('quantity', 0)))
        notes = body.get('notes', '')

        # Get product
        product = get_object_or_404(Seafood, id=product_id)
        old_quantity = product.stock_quantity

        # Calculate new quantity based on type
        if adjustment_type == 'import':
            # Nhập hàng: luôn cộng thêm
            weight_change = abs(quantity)
            new_quantity = old_quantity + weight_change
        elif adjustment_type == 'loss':
            # Hao hụt: luôn trừ đi
            weight_change = -abs(quantity)
            new_quantity = old_quantity + weight_change
        else:
            # Điều chỉnh: có thể cộng hoặc trừ
            weight_change = quantity
            new_quantity = old_quantity + weight_change

        # Validate
        if new_quantity < 0:
            return JsonResponse({
                "success": False,
                "error": f"Số lượng không thể âm. Hiện tại: {old_quantity} kg, Thay đổi: {weight_change} kg"
            }, status=400)

        # Update product stock
        product.stock_quantity = new_quantity
        product.save()

        # Create inventory log
        InventoryLog.objects.create(
            seafood=product,
            type=adjustment_type,
            weight_change=weight_change,
            stock_after=new_quantity,
            notes=notes,
            created_by=None  # TODO: Add user from auth
        )

        return {
            "success": True,
            "old_quantity": float(old_quantity),
            "new_quantity": float(new_quantity),
            "change": float(weight_change),
            "type": adjustment_type
        }

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


@router.get("/products/{product_id}/inventory-logs")
def get_inventory_logs(request, product_id: UUID, limit: int = 20):
    """Lấy lịch sử điều chỉnh kho của sản phẩm"""
    product = get_object_or_404(Seafood, id=product_id)

    logs = InventoryLog.objects.filter(
        seafood=product
    ).order_by('-created_at')[:limit]

    return [
        {
            "id": str(log.id),
            "type": log.type,
            "weight_change": float(log.weight_change),
            "stock_after": float(log.stock_after),
            "notes": log.notes,
            "created_at": log.created_at.isoformat(),
            "created_by": log.created_by.username if log.created_by else None
        }
        for log in logs
    ]


# ============================================
# IMPORT SOURCE ENDPOINTS
# ============================================

@router.get("/import-sources", response=List[ImportSourceRead])
def list_import_sources(request):
    """Lấy danh sách nguồn nhập hàng"""
    return ImportSource.objects.filter(is_active=True).all()


@router.post("/import-sources", response=ImportSourceRead)
def create_import_source(request, payload: ImportSourceCreate):
    """Tạo nguồn nhập hàng mới"""
    source = ImportSource.objects.create(**payload.dict())
    return source


@router.put("/import-sources/{source_id}", response=ImportSourceRead)
def update_import_source(request, source_id: UUID, payload: ImportSourceUpdate):
    """Cập nhật nguồn nhập hàng"""
    source = get_object_or_404(ImportSource, id=source_id)
    for key, value in payload.dict(exclude_unset=True).items():
        setattr(source, key, value)
    source.save()
    return source


# ============================================
# IMPORT BATCH ENDPOINTS
# ============================================

@router.get("/import-batches", response=List[ImportBatchRead])
def list_import_batches(request, seafood_id: UUID = None):
    """Lấy danh sách lô nhập hàng"""
    query = ImportBatch.objects.filter(is_active=True).select_related(
        'seafood', 'seafood__category', 'import_source'
    )

    if seafood_id:
        query = query.filter(seafood_id=seafood_id)

    return query.all()


@router.get("/import-batches/{batch_id}", response=ImportBatchRead)
def get_import_batch(request, batch_id: UUID):
    """Lấy chi tiết lô nhập hàng"""
    return get_object_or_404(
        ImportBatch.objects.select_related('seafood', 'seafood__category', 'import_source'),
        id=batch_id
    )


@router.post("/import-batches", response=ImportBatchRead)
def create_import_batch(request, payload: ImportBatchCreate):
    """Tạo lô nhập hàng mới"""
    data = payload.dict()

    # Tạo batch_code tự động nếu không có
    if not data.get('batch_code'):
        today = timezone.now().strftime('%Y%m%d')
        count = ImportBatch.objects.filter(batch_code__startswith=f'IMP-{today}').count()
        data['batch_code'] = f'IMP-{today}-{count + 1:03d}'

    # Get user from request or use first available user
    from django.contrib.auth import get_user_model
    User = get_user_model()

    if hasattr(request, 'user') and request.user.is_authenticated:
        data['imported_by'] = request.user
    else:
        data['imported_by'] = User.objects.first()

    batch = ImportBatch.objects.create(**data)

    # Cập nhật stock sản phẩm
    seafood = batch.seafood
    seafood.stock_quantity = Decimal(str(seafood.stock_quantity)) + batch.total_weight
    seafood.current_price = batch.sell_price
    seafood.save()

    # Tạo inventory log
    InventoryLog.objects.create(
        seafood=seafood,
        import_batch=batch,
        type='import',
        weight_change=batch.total_weight,
        stock_after=seafood.stock_quantity,
        notes=f'Nhập lô {batch.batch_code}',
        created_by=data['imported_by']
    )

    return batch


@router.put("/import-batches/{batch_id}", response=ImportBatchRead)
def update_import_batch(request, batch_id: UUID, payload: ImportBatchUpdate):
    """Cập nhật lô nhập hàng"""
    batch = get_object_or_404(ImportBatch, id=batch_id)
    for key, value in payload.dict(exclude_unset=True).items():
        setattr(batch, key, value)
    batch.save()
    return batch


# ============================================
# ORDER ENDPOINTS (POS)
# ============================================

@router.get("/orders", response=List[OrderRead])
def list_orders(request, status: str = None, customer_phone: str = None, created_by: UUID = None, limit: int = 50):
    """Lấy danh sách đơn hàng"""
    query = Order.objects.filter(is_active=True).prefetch_related(
        'items', 'items__seafood', 'items__seafood__category'
    )

    if status:
        query = query.filter(status=status)
    if customer_phone:
        query = query.filter(customer_phone__icontains=customer_phone)
    if created_by:
        query = query.filter(created_by_id=created_by)

    orders = query.order_by('-created_at')[:limit]

    # Convert to list of dicts to avoid RelatedManager serialization issue
    result = []
    for order in orders:
        items_list = list(order.items.all())
        order_dict = {
            'id': order.id,
            'order_code': order.order_code,
            'customer_name': order.customer_name or '',
            'customer_phone': order.customer_phone,
            'customer_address': order.customer_address or '',
            'payment_method': order.payment_method or '',
            'payment_status': order.payment_status,
            'status': order.status,
            'notes': order.notes or '',
            'discount_amount': order.discount_amount,
            'subtotal': order.subtotal,
            'total_amount': order.total_amount,
            'paid_amount': order.paid_amount,
            'created_at': order.created_at,
            'weighed_at': order.weighed_at,
            'weighed_by': order.weighed_by.email if order.weighed_by else None,
            'weight_images': order.weight_images,
            'shipped_at': order.shipped_at,
            'shipped_by': order.shipped_by.email if order.shipped_by else None,
            'shipping_notes': order.shipping_notes or '',
            'items': items_list
        }
        result.append(order_dict)

    return result


@router.get("/orders/{order_id}", response=OrderRead)
def get_order(request, order_id: UUID):
    """Lấy chi tiết đơn hàng"""
    order = get_object_or_404(
        Order.objects.prefetch_related('items', 'items__seafood', 'items__seafood__category'),
        id=order_id
    )
    # Serialize items manually to avoid OrderItem model object serialization issues
    items_list = [
        {
            'id': str(item.id),
            'seafood_id': str(item.seafood_id),
            'quantity': float(item.quantity) if item.quantity else None,
            'weight': float(item.weight),
            'unit_price': float(item.unit_price),
            'subtotal': float(item.subtotal) if hasattr(item, 'subtotal') else float(item.weight * item.unit_price),
            'estimated_weight': float(item.estimated_weight) if item.estimated_weight else None,
            'weight_image_url': item.weight_image_url or '',
            'notes': item.notes or '',
            'seafood': {
                'id': str(item.seafood.id),
                'code': item.seafood.code,
                'name': item.seafood.name,
                'description': item.seafood.description or '',
                'unit_type': item.seafood.unit_type,
                'avg_unit_weight': float(item.seafood.avg_unit_weight) if item.seafood.avg_unit_weight else None,
                'current_price': float(item.seafood.current_price),
                'stock_quantity': float(item.seafood.stock_quantity) if item.seafood.stock_quantity else 0,
                'status': item.seafood.status,
                'created_at': item.seafood.created_at,
                'is_active': item.seafood.status == 'active',
                'category_id': str(item.seafood.category_id) if item.seafood.category_id else None,
                'category': {
                    'id': str(item.seafood.category.id),
                    'name': item.seafood.category.name,
                    'slug': item.seafood.category.slug,
                    'description': item.seafood.category.description or '',
                    'image_url': item.seafood.category.image_url or '',
                    'sort_order': item.seafood.category.sort_order,
                    'created_at': item.seafood.category.created_at,
                    'is_active': item.seafood.category.is_active,
                } if item.seafood.category else None,
            } if item.seafood else None,
        }
        for item in order.items.all()
    ]
    return {
        'id': order.id,
        'order_code': order.order_code,
        'customer_name': order.customer_name or '',
        'customer_phone': order.customer_phone,
        'customer_address': order.customer_address or '',
        'customer_source': order.customer_source or '',
        'payment_method': order.payment_method or '',
        'payment_status': order.payment_status,
        'status': order.status,
        'notes': order.notes or '',
        'discount_amount': float(order.discount_amount),
        'subtotal': float(order.subtotal),
        'total_amount': float(order.total_amount),
        'paid_amount': float(order.paid_amount),
        'created_at': order.created_at,
        'customer_id': order.customer_id,
        'sale_user_id': order.sale_user_id,
        'confirmed_by_sale_at': order.confirmed_by_sale_at,
        'assigned_employee_id': order.assigned_employee_id,
        'assigned_at': order.assigned_at,
        'weighed_at': order.weighed_at,
        'weighed_by_id': order.weighed_by_id,
        'weight_images': order.weight_images or [],
        'shipped_at': order.shipped_at,
        'shipped_by_id': order.shipped_by_id,
        'shipping_notes': order.shipping_notes or '',
        'items': items_list
    }


@router.post("/orders", response=OrderRead)
def create_order(request, payload: OrderCreate):
    """Tạo đơn hàng POS mới"""
    data = payload.dict()
    items_data = data.pop('items')

    # Tạo order_code tự động
    today = timezone.now().strftime('%Y%m%d')
    count = Order.objects.filter(order_code__startswith=f'POS-{today}').count()
    data['order_code'] = f'POS-{today}-{count + 1:03d}'

    # Tính tổng tiền
    subtotal = Decimal('0')
    for item in items_data:
        item_total = Decimal(str(item['weight'])) * Decimal(str(item['unit_price']))
        subtotal += item_total

    data['subtotal'] = subtotal
    data['total_amount'] = subtotal - Decimal(str(data.get('discount_amount', 0)))
    data['paid_amount'] = data['total_amount'] if data.get('payment_status') == 'paid' else Decimal('0')

    # Get user from request or use first available user
    from django.contrib.auth import get_user_model
    User = get_user_model()

    # Try to get authenticated user from request
    if hasattr(request, 'auth') and request.auth:
        user = request.auth

        # If user is a customer, link the order to their account
        if user.user_type == 'customer':
            data['customer'] = user
            data['status'] = 'pending_sale_confirm'  # Customer orders need Sale confirmation
            # Auto-fill customer info from their profile if not provided
            if not data.get('customer_name'):
                data['customer_name'] = user.full_name
            if not data.get('customer_phone'):
                data['customer_phone'] = user.phone or ''

        # If user is sale/employee, they're creating order for walk-in customer
        elif user.user_type in ['employee', 'manager']:
            data['status'] = 'pending'  # Sale orders go directly to pending
            # Check if user has sale role or is manager
            if user.has_role('sale') or user.user_type == 'manager':
                data['sale_user'] = user

        # Set created_by to the authenticated user
        data['created_by'] = user
    elif hasattr(request, 'user') and request.user.is_authenticated:
        data['created_by'] = request.user
    else:
        # Fallback: get any user or create default one
        user = User.objects.first()
        if not user:
            # Create a default system user
            user = User.objects.create_user(
                email='system@seafood.com',
                password='system123',
                first_name='System',
                last_name='User'
            )
        data['created_by'] = user

    # Tạo order
    order = Order.objects.create(**data)

    # Tạo order items và cập nhật stock
    for item_data in items_data:
        seafood = Seafood.objects.get(id=item_data['seafood_id'])
        weight = Decimal(str(item_data['weight']))

        # Tạo order item
        order_item = OrderItem.objects.create(
            order=order,
            seafood=seafood,
            import_batch_id=item_data.get('import_batch_id'),
            weight=weight,
            unit_price=item_data['unit_price'],
            subtotal=weight * Decimal(str(item_data['unit_price'])),
            notes=item_data.get('notes', '')
        )

        # Cập nhật stock sản phẩm
        seafood.stock_quantity = Decimal(str(seafood.stock_quantity)) - weight
        if seafood.stock_quantity <= 0:
            seafood.status = 'out_of_stock'
        seafood.save()

        # Cập nhật batch remaining weight
        if item_data.get('import_batch_id'):
            batch = ImportBatch.objects.get(id=item_data['import_batch_id'])
            batch.remaining_weight = Decimal(str(batch.remaining_weight)) - weight
            if batch.remaining_weight <= 0:
                batch.status = 'sold_out'
            batch.save()

        # Tạo inventory log
        InventoryLog.objects.create(
            seafood=seafood,
            import_batch_id=item_data.get('import_batch_id'),
            order_item=order_item,
            type='sale',
            weight_change=-weight,
            stock_after=seafood.stock_quantity,
            notes=f'Bán cho {order.customer_phone}',
            created_by=data['created_by']
        )

    return get_order(request, order.id)


@router.put("/orders/{order_id}", response=OrderRead)
def update_order(request, order_id: UUID, payload: OrderUpdate):
    """Cập nhật đơn hàng"""
    order = get_object_or_404(Order, id=order_id)
    for key, value in payload.dict(exclude_unset=True).items():
        setattr(order, key, value)

    # Cập nhật paid_amount nếu đổi payment_status
    if hasattr(payload, 'payment_status') and payload.payment_status == 'paid':
        order.paid_amount = order.total_amount

    order.save()
    return get_order(request, order_id)


@router.delete("/orders/{order_id}")
def cancel_order(request, order_id: UUID):
    """Hủy đơn hàng"""
    order = get_object_or_404(Order, id=order_id)
    order.status = 'cancelled'
    order.save()

    # Hoàn lại stock
    for item in order.items.all():
        seafood = item.seafood
        weight = Decimal(str(item.weight))
        seafood.stock_quantity = Decimal(str(seafood.stock_quantity)) + weight
        seafood.save()

        # Hoàn lại batch
        if item.import_batch:
            batch = item.import_batch
            batch.remaining_weight = Decimal(str(batch.remaining_weight)) + weight
            batch.save()

    return {"success": True}


# ============================================
# ORDER WORKFLOW ENDPOINTS
# ============================================

@router.post("/orders/{order_id}/update-item")
def update_order_item(
    request,
    order_id: UUID,
    item_id: str = None,
    weight: float = None,
    unit_price: int = None,
    weight_image_url: str = None
):
    """Cập nhật trọng lượng và giá của order item (sau khi cân)"""
    order = get_object_or_404(Order, id=order_id)

    if not item_id:
        return {"success": False, "error": "item_id is required"}

    # Convert string to UUID
    try:
        item_id_uuid = UUID(item_id)
    except (ValueError, AttributeError):
        return {"success": False, "error": "Invalid item_id format"}

    item = get_object_or_404(OrderItem, id=item_id_uuid, order=order)

    # Lưu trọng lượng cũ để tính chênh lệch stock
    old_weight = item.weight

    # Cập nhật item
    if weight is not None:
        item.weight = Decimal(str(weight))
    if unit_price is not None:
        item.unit_price = Decimal(str(unit_price))
    if weight_image_url is not None:
        item.weight_image_url = weight_image_url

    # Tính lại subtotal
    item.subtotal = item.weight * item.unit_price
    item.save()

    # Cập nhật stock nếu thay đổi trọng lượng
    if weight is not None:
        weight_diff = item.weight - old_weight
        seafood = item.seafood
        seafood.stock_quantity = Decimal(str(seafood.stock_quantity)) - weight_diff
        seafood.save()

        # Cập nhật batch
        if item.import_batch:
            batch = item.import_batch
            batch.remaining_weight = Decimal(str(batch.remaining_weight)) - weight_diff
            batch.save()

    # Tính lại tổng tiền đơn hàng
    order.subtotal = sum(
        Decimal(str(i.subtotal)) for i in order.items.all()
    )
    order.total_amount = order.subtotal - order.discount_amount
    if order.payment_status == 'paid':
        order.paid_amount = order.total_amount
    order.save()

    return {
        "success": True,
        "item": item,
        "order": order
    }


class MarkWeighedSchema(BaseModel):
    weight_images: List[str] = []

class MarkShippedSchema(BaseModel):
    shipping_notes: str = ""

@router.post("/orders/{order_id}/items/{item_id}/upload-image")
def upload_weight_image(request, order_id: UUID, item_id: UUID):
    """Upload ảnh cân cho order item"""
    from django.core.files.storage import default_storage
    from django.core.files.base import ContentFile
    import os
    from datetime import datetime

    order = get_object_or_404(Order, id=order_id)
    item = get_object_or_404(OrderItem, id=item_id, order=order)

    # Get uploaded file from request
    if not request.FILES or 'image' not in request.FILES:
        return {"success": False, "error": "No image file provided"}

    image_file = request.FILES['image']

    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    ext = os.path.splitext(image_file.name)[1]
    filename = f'weight_images/{order.order_code}_{item_id}_{timestamp}{ext}'

    # Save file
    path = default_storage.save(filename, ContentFile(image_file.read()))

    # Build full URL
    request_host = request.META.get('HTTP_HOST', 'localhost:8003')
    full_url = f"http://{request_host}/media/{path}"

    # Update item
    item.weight_image_url = full_url
    item.save()

    return {"success": True, "image_url": full_url}

@router.post("/orders/{order_id}/mark-paid")
def mark_order_paid(request, order_id: UUID):
    """Staff đánh dấu đã thu tiền - chuyển sang chờ xác minh"""
    order = get_object_or_404(Order, id=order_id)
    order.payment_status = 'pending_verification'  # Chờ Sale xác minh
    order.save()
    return {"success": True, "message": "Đã đánh dấu đã thu tiền, chờ Sale xác minh"}

@router.post("/orders/{order_id}/verify-payment")
def verify_payment(request, order_id: UUID):
    """Sale xác minh thanh toán"""
    order = get_object_or_404(Order, id=order_id)
    order.payment_status = 'paid'
    order.paid_amount = order.total_amount
    order.save()
    return {"success": True, "message": "Đã xác minh thanh toán"}

@router.post("/orders/{order_id}/mark-weighed")
def mark_order_weighed(request, order_id: UUID, payload: MarkWeighedSchema = None):
    """Đánh dấu đơn hàng đã cân xong"""
    order = get_object_or_404(Order, id=order_id)

    order.status = 'weighed'
    order.weighed_at = timezone.now()

    # Get current user (for now, use admin)
    from django.contrib.auth import get_user_model
    User = get_user_model()
    order.weighed_by = User.objects.filter(is_superuser=True).first()

    if payload and payload.weight_images:
        order.weight_images = payload.weight_images

    order.save()

    return {
        "success": True,
        "order": get_order(request, order_id)
    }


@router.post("/orders/{order_id}/mark-shipped")
def mark_order_shipped(request, order_id: UUID, payload: MarkShippedSchema = None):
    """Đánh dấu đơn hàng đã gửi vận chuyển"""
    order = get_object_or_404(Order, id=order_id)

    order.status = 'shipped'
    order.shipped_at = timezone.now()

    # Get current user (for now, use admin)
    from django.contrib.auth import get_user_model
    User = get_user_model()
    order.shipped_by = User.objects.filter(is_superuser=True).first()

    if payload and payload.shipping_notes:
        order.shipping_notes = payload.shipping_notes

    order.save()

    return {
        "success": True,
        "order": get_order(request, order_id)
    }


@router.post("/orders/{order_id}/mark-delivered")
def mark_order_delivered(request, order_id: UUID):
    """Xác nhận đơn hàng đã giao thành công"""
    order = get_object_or_404(Order, id=order_id)

    if order.status != 'shipped':
        from api.exceptions import BadRequest
        raise BadRequest("Chỉ có thể xác nhận giao hàng cho đơn đã gửi vận chuyển")

    order.status = 'completed'
    order.delivered_at = timezone.now()

    # Get current user from request
    if hasattr(request, 'auth') and request.auth:
        order.delivered_by = request.auth

    order.save()

    return {
        "success": True,
        "message": "Đã xác nhận giao hàng thành công"
    }


@router.get("/orders/{order_id}/export-pdf")
def export_order_pdf(request, order_id: UUID):
    """Xuất PDF hóa đơn"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO

    order = get_object_or_404(
        Order.objects.prefetch_related('items', 'items__seafood'),
        id=order_id
    )

    # Tạo PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("HÓA ĐƠN BÁN HÀNG", title_style))
    elements.append(Spacer(1, 20))

    # Thông tin đơn hàng
    info_data = [
        ['Mã đơn hàng:', order.order_code],
        ['Ngày tạo:', order.created_at.strftime('%d/%m/%Y %H:%M')],
        ['Khách hàng:', order.customer_name or 'N/A'],
        ['Số điện thoại:', order.customer_phone],
        ['Địa chỉ:', order.customer_address or 'N/A'],
        ['Trạng thái:', order.get_status_display()],
    ]

    info_table = Table(info_data, colWidths=[4*cm, 12*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 30))

    # Bảng sản phẩm
    items_data = [['STT', 'Sản phẩm', 'Số lượng (kg)', 'Đơn giá', 'Thành tiền']]

    for idx, item in enumerate(order.items.all(), 1):
        items_data.append([
            str(idx),
            item.seafood.name,
            f"{item.weight:,.2f}",
            f"{item.unit_price:,.0f}đ",
            f"{item.subtotal:,.0f}đ"
        ])

    items_table = Table(items_data, colWidths=[1.5*cm, 7*cm, 3*cm, 3.5*cm, 4*cm])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 30))

    # Tổng tiền
    total_data = [
        ['Tạm tính:', f"{order.subtotal:,.0f}đ"],
        ['Giảm giá:', f"-{order.discount_amount:,.0f}đ"],
        ['TỔNG CỘNG:', f"{order.total_amount:,.0f}đ"],
    ]

    total_table = Table(total_data, colWidths=[14*cm, 5*cm])
    total_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 14),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#1e40af')),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#1e40af')),
        ('TOPPADDING', (0, -1), (-1, -1), 15),
    ]))
    elements.append(total_table)

    # Build PDF
    doc.build(elements)

    # Trả về response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="order_{order.order_code}.pdf"'

    return response


# ============================================
# DASHBOARD / STATS ENDPOINTS
# ============================================

@router.get("/stats/dashboard", response=DashboardStats)
def get_dashboard_stats(request):
    """Lấy thống kê tổng quan"""
    today = timezone.now().date()

    # Tổng số sản phẩm
    total_products = Seafood.objects.filter(is_active=True, status='active').count()

    # Tổng giá trị tồn kho
    total_stock_value = Seafood.objects.filter(is_active=True).aggregate(
        total=Sum(F('stock_quantity') * F('current_price'))
    )['total'] or Decimal('0')

    # Đơn hàng hôm nay
    today_orders = Order.objects.filter(
        created_at__date=today,
        status__in=['pending', 'completed']
    ).count()

    # Doanh thu hôm nay
    today_revenue = Order.objects.filter(
        created_at__date=today,
        status='completed'
    ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')

    # Sản phẩm sắp hết hàng (< 10kg)
    low_stock_products = Seafood.objects.filter(
        is_active=True,
        stock_quantity__lt=10,
        status='active'
    ).count()

    return DashboardStats(
        total_products=total_products,
        total_stock_value=total_stock_value,
        today_orders=today_orders,
        today_revenue=today_revenue,
        low_stock_products=low_stock_products
    )


@router.get("/stats/products", response=List[ProductStats])
def get_product_stats(request, limit: int = 10):
    """Lấy thống kê sản phẩm bán chạy"""
    products = Seafood.objects.filter(is_active=True).annotate(
        total_sold=Sum('orderitem__weight'),
        revenue=Sum(F('orderitem__weight') * F('orderitem__unit_price'))
    ).order_by('-revenue')[:limit]

    return [
        ProductStats(
            code=p.code,
            name=p.name,
            stock_quantity=p.stock_quantity,
            current_price=p.current_price,
            total_sold=p.total_sold or Decimal('0'),
            revenue=p.revenue or Decimal('0')
        )
        for p in products
    ]


# ============================================
# ORDER WORKFLOW APIs
# ============================================

@router.post("/orders/{order_id}/confirm-by-sale", response=OrderRead)
def confirm_order_by_sale(request, order_id: UUID, payload: OrderConfirmBySale):
    """Sale xác nhận đơn hàng từ customer"""
    order = get_object_or_404(Order, id=order_id)
    
    # Check order status
    if order.status != 'pending_sale_confirm':
        return {"error": "Order is not pending sale confirmation"}
    
    # Update order
    order.status = 'confirmed'
    order.confirmed_by_sale_at = timezone.now()
    
    # Set sale_user if authenticated
    if hasattr(request, 'auth') and request.auth:
        order.sale_user = request.auth
    
    if payload.notes:
        order.notes = (order.notes or '') + f"\n[Sale confirmed]: {payload.notes}"
    
    order.save()
    
    return get_order(request, order_id)


@router.post("/orders/{order_id}/assign-to-employee", response=OrderRead)
def assign_order_to_employee(request, order_id: UUID, payload: OrderAssignToEmployee):
    """Sale giao đơn hàng cho nhân viên kho"""
    order = get_object_or_404(Order, id=order_id)
    
    # Check if order can be assigned
    if order.status not in ['pending', 'confirmed']:
        return {"error": "Order cannot be assigned in current status"}
    
    # Get employee
    from django.contrib.auth import get_user_model
    User = get_user_model()
    employee = get_object_or_404(User, id=payload.employee_id)
    
    # Update order
    order.assigned_employee = employee
    order.assigned_at = timezone.now()
    order.status = 'assigned_to_warehouse'
    
    if payload.notes:
        order.notes = (order.notes or '') + f"\n[Assigned to {employee.full_name}]: {payload.notes}"
    
    order.save()
    
    return get_order(request, order_id)


@router.post("/orders/{order_id}/start-weighing", response=OrderRead)
def start_weighing_order(request, order_id: UUID, payload: OrderStartWeighing):
    """Employee bắt đầu cân hàng"""
    order = get_object_or_404(Order, id=order_id)
    
    # Check if order can be weighed
    if order.status != 'assigned_to_warehouse':
        return {"error": "Order is not ready for weighing"}
    
    # Update order
    order.status = 'weighing'
    
    if payload.notes:
        order.notes = (order.notes or '') + f"\n[Weighing started]: {payload.notes}"
    
    order.save()
    
    return get_order(request, order_id)


@router.post("/orders/{order_id}/complete-weighing", response=OrderRead)
def complete_weighing_order(request, order_id: UUID, payload: OrderCompleteWeighing):
    """Employee hoàn thành cân hàng, cập nhật kg/giá và upload ảnh"""
    order = get_object_or_404(Order, id=order_id)
    
    # Check if order is being weighed
    if order.status != 'weighing':
        return {"error": "Order is not in weighing status"}
    
    # Update order items with actual weight and price
    for item_update in payload.item_updates:
        item = get_object_or_404(OrderItem, id=item_update.item_id, order=order)
        
        # Update weight
        old_weight = item.weight
        item.weight = item_update.actual_weight
        
        # Update price if provided
        if item_update.actual_unit_price:
            item.unit_price = item_update.actual_unit_price
        
        # Recalculate subtotal
        item.subtotal = item.weight * item.unit_price
        
        # Add notes
        if item_update.notes:
            item.notes = (item.notes or '') + f"\n[Weight updated from {old_weight}kg to {item.weight}kg]: {item_update.notes}"
        
        item.save()
        
        # Update stock (adjust difference)
        seafood = item.seafood
        weight_diff = item.weight - old_weight
        seafood.stock_quantity -= weight_diff
        if seafood.stock_quantity <= 0:
            seafood.status = 'out_of_stock'
        seafood.save()
    
    # Recalculate order totals
    order.subtotal = sum(Decimal(str(i.subtotal)) for i in order.items.all())
    order.total_amount = order.subtotal - order.discount_amount
    if order.payment_status == 'paid':
        order.paid_amount = order.total_amount
    
    # Update order status and images
    order.status = 'weighed'
    order.weighed_at = timezone.now()
    order.weight_images = payload.weight_images
    
    # Set weighed_by if authenticated
    if hasattr(request, 'auth') and request.auth:
        order.weighed_by = request.auth
    
    if payload.notes:
        order.notes = (order.notes or '') + f"\n[Weighing completed]: {payload.notes}"
    
    order.save()
    
    return get_order(request, order_id)


@router.get("/orders/by-role/sale", response=List[OrderRead])
def get_sale_orders(request, status: str = None):
    """Lấy danh sách đơn hàng cho Sale - Xem TẤT CẢ đơn hàng"""
    # Sale có thể xem TẤT CẢ đơn hàng (không filter theo sale_user)
    # Bao gồm: đơn cần xác nhận, đơn đang xử lý, đơn đã cân
    orders = Order.objects.exclude(
        status__in=['cancelled']
    ).select_related('created_by', 'sale_user', 'assigned_employee', 'customer', 'weighed_by', 'shipped_by').prefetch_related('items__seafood').order_by('-created_at')

    if status:
        orders = orders.filter(status=status)

    # Convert to list of dicts to avoid RelatedManager serialization issue
    result = []
    for order in orders:
        items_list = list(order.items.all())
        order_dict = {
            'id': order.id,
            'order_code': order.order_code,
            'customer_name': order.customer_name or '',
            'customer_phone': order.customer_phone,
            'customer_address': order.customer_address or '',
            'customer_source': order.customer_source or '',
            'payment_method': order.payment_method or '',
            'payment_status': order.payment_status,
            'status': order.status,
            'notes': order.notes or '',
            'discount_amount': order.discount_amount,
            'subtotal': order.subtotal,
            'total_amount': order.total_amount,
            'paid_amount': order.paid_amount,
            'created_at': order.created_at,
            'customer_id': order.customer_id,
            'sale_user_id': order.sale_user_id,
            'confirmed_by_sale_at': order.confirmed_by_sale_at,
            'assigned_employee_id': order.assigned_employee_id,
            'assigned_at': order.assigned_at,
            'weighed_at': order.weighed_at,
            'weighed_by_id': order.weighed_by_id,
            'weight_images': order.weight_images,
            'shipped_at': order.shipped_at,
            'shipped_by_id': order.shipped_by_id,
            'shipping_notes': order.shipping_notes or '',
            'items': items_list
        }
        result.append(order_dict)

    return result


@router.get("/orders/by-role/employee", response=List[OrderRead])
def get_employee_orders(request, status: str = None):
    """Lấy danh sách đơn hàng cho Employee (nhân viên kho) - CHỈ đơn được giao"""
    # Employee CHỈ thấy đơn được giao cho mình
    query = Q(status='assigned_to_warehouse') | Q(status='weighing') | Q(status='weighed')

    if hasattr(request, 'auth') and request.auth:
        # PHẢI filter theo assigned_employee
        query = query & Q(assigned_employee=request.auth)
    else:
        # Nếu không đăng nhập, không trả về gì
        return []

    orders = Order.objects.filter(query).select_related('created_by', 'sale_user', 'assigned_employee', 'customer', 'weighed_by', 'shipped_by').prefetch_related('items__seafood').order_by('-created_at')

    if status:
        orders = orders.filter(status=status)

    # Convert to list of dicts to avoid RelatedManager serialization issue
    result = []
    for order in orders:
        items_list = list(order.items.all())
        order_dict = {
            'id': order.id,
            'order_code': order.order_code,
            'customer_name': order.customer_name or '',
            'customer_phone': order.customer_phone,
            'customer_address': order.customer_address or '',
            'customer_source': order.customer_source or '',
            'payment_method': order.payment_method or '',
            'payment_status': order.payment_status,
            'status': order.status,
            'notes': order.notes or '',
            'discount_amount': order.discount_amount,
            'subtotal': order.subtotal,
            'total_amount': order.total_amount,
            'paid_amount': order.paid_amount,
            'created_at': order.created_at,
            'customer_id': order.customer_id,
            'sale_user_id': order.sale_user_id,
            'confirmed_by_sale_at': order.confirmed_by_sale_at,
            'assigned_employee_id': order.assigned_employee_id,
            'assigned_at': order.assigned_at,
            'weighed_at': order.weighed_at,
            'weighed_by_id': order.weighed_by_id,
            'weight_images': order.weight_images,
            'shipped_at': order.shipped_at,
            'shipped_by_id': order.shipped_by_id,
            'shipping_notes': order.shipping_notes or '',
            'items': items_list
        }
        result.append(order_dict)

    return result


# ===========================
# SePay Payment Webhook
# ===========================

@router.post("/sepay/webhook", auth=None)
def sepay_webhook(request):
    """
    SePay webhook callback for payment verification
    This endpoint receives payment notifications from SePay

    Expected payload from SePay:
    {
        "gateway": "SePay",
        "transaction_id": "xxx",
        "reference_number": "ORDER123",  // Our order code
        "amount": 1000000,
        "content": "Thanh toan ORDER123",
        "sub_account": "123456789",  // Bank account number
        "bank_code": "ACB",
        "status": "success"  // or "pending", "failed"
    }
    """
    try:
        import json

        # Parse request body
        if request.body:
            payload = json.loads(request.body)
        else:
            return {"success": False, "error": "Empty payload"}

        # Extract data from SePay payload
        reference_number = payload.get('reference_number', '')
        amount = float(payload.get('amount', 0))
        status = payload.get('status', '')
        transaction_id = payload.get('transaction_id', '')
        content = payload.get('content', '')

        # Find order by order_code in reference_number or content
        order_code = reference_number
        if not order_code:
            # Try to extract order code from content
            # Content format: "Thanh toan ORDER123"
            import re
            match = re.search(r'ORD[A-Z0-9]+', content.upper())
            if match:
                order_code = match.group(0)

        if not order_code:
            return {"success": False, "error": "Order code not found in payload"}

        # Find the order
        try:
            order = Order.objects.get(order_code=order_code)
        except Order.DoesNotExist:
            return {"success": False, "error": f"Order {order_code} not found"}

        # Verify amount matches (allow 1% difference for rounding)
        expected_amount = float(order.total_amount)
        if abs(amount - expected_amount) > (expected_amount * 0.01):
            return {
                "success": False,
                "error": f"Amount mismatch. Expected: {expected_amount}, Received: {amount}"
            }

        # Update payment status based on SePay status
        if status == 'success':
            order.payment_status = 'paid'
            order.paid_amount = amount

            # Store transaction info in notes if needed
            transaction_note = f"\nGiao dịch SePay: {transaction_id} - {amount}đ"
            if order.notes:
                order.notes += transaction_note
            else:
                order.notes = transaction_note.strip()

            order.save()

            return {
                "success": True,
                "message": f"Payment confirmed for order {order_code}",
                "order_id": str(order.id)
            }
        elif status == 'pending':
            order.payment_status = 'pending_verification'
            order.save()
            return {
                "success": True,
                "message": f"Payment pending for order {order_code}"
            }
        else:
            # Payment failed
            return {
                "success": False,
                "error": f"Payment failed with status: {status}"
            }

    except Exception as e:
        import traceback
        return {
            "success": False,
            "error": str(e),
            "traceback": traceback.format_exc()
        }


@router.get("/payment/check-order/{order_code}")
def check_order_payment(request, order_code: str):
    """
    Check payment status of an order
    This can be called from frontend to poll payment status
    """
    try:
        order = Order.objects.get(order_code=order_code)
        return {
            "success": True,
            "order_code": order.order_code,
            "payment_status": order.payment_status,
            "payment_method": order.payment_method,
            "total_amount": float(order.total_amount),
            "paid_amount": float(order.paid_amount),
            "is_paid": order.payment_status == 'paid'
        }
    except Order.DoesNotExist:
        return {"success": False, "error": "Order not found"}


@router.post("/payment/generate-qr/{order_code}")
def generate_payment_qr(request, order_code: str):
    """
    Generate QR code for payment using SePay API
    Returns QR code image URL and payment details

    IMPORTANT: QR code can only be generated AFTER order has been weighed
    """
    try:
        order = Order.objects.get(order_code=order_code)

        # Check if order has been weighed - ONLY allow QR generation after weighing
        if order.status not in ['weighed', 'completed']:
            return {
                "success": False,
                "error": "Đơn hàng chưa được cân. Vui lòng chờ nhân viên cân hàng và gửi ảnh trước khi thanh toán.",
                "status": order.status,
                "can_pay": False
            }

        # Check if already paid
        if order.payment_status == 'paid':
            return {
                "success": False,
                "error": "Đơn hàng đã được thanh toán",
                "can_pay": False
            }

        # Get SePay service
        sepay = get_sepay_service()

        # Generate QR code
        content = f"Thanh toan {order_code}"
        qr_data = sepay.create_qr_code(
            amount=order.total_amount,
            content=content
        )

        return {
            "success": True,
            "order_code": order_code,
            "amount": float(order.total_amount),
            "qr_image_url": qr_data['qr_image_url'],
            "qr_svg": qr_data.get('qr_svg', ''),
            "account_number": qr_data['account_number'],
            "account_name": qr_data['account_name'],
            "bank_code": qr_data['bank_code'],
            "content": content,
            "session_id": qr_data.get('session_id', ''),
            "can_pay": True
        }
    except Order.DoesNotExist:
        return {"success": False, "error": "Order not found", "can_pay": False}
    except SepayAPIError as e:
        return {"success": False, "error": str(e), "can_pay": False}
